from django.contrib import admin
from django.db.models import Sum, F
from django.utils.html import format_html
from .models import *
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


class PurchaseOrderLineItemInline(admin.TabularInline):
    model = PurchaseOrderLineItem
    extra = 1


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'sku', 'unit_price')
    search_fields = ('name', 'sku')

@admin.register(PurchaseOrder)
class PurchaseOrderAdmin(admin.ModelAdmin):
    inlines = [PurchaseOrderLineItemInline]
    list_display = ('id', 'vendor', 'order_date', 'status', 'total_cost_display')
    
    def total_cost_display(self, obj):
        return f"${obj.total_cost:.2f}"
    total_cost_display.short_description = 'Total Cost'

class InvoiceLineItemInline(admin.TabularInline):
    model = InvoiceLineItem
    extra = 1

@admin.register(Invoice)
class InvoiceAdmin(admin.ModelAdmin):
    inlines = [InvoiceLineItemInline]
    list_display = ('id', 'customer', 'invoice_date', 'due_date', 'status', 
                   'total_amount_display', 'is_overdue')
    actions = ['mark_as_paid', 'export_to_xlsx']
    
    # Custom list columns
    def total_amount_display(self, obj):
        return f"${obj.total_amount:.2f}"
    total_amount_display.short_description = 'Total Amount'
    
    def is_overdue(self, obj):
        if obj.status == 'unpaid' and obj.due_date < timezone.now().date():
            return format_html('<span style="color: red;">Overdue</span>')
        return ''
    is_overdue.short_description = 'Status'
    
    # Custom actions
    def mark_as_paid(self, request, queryset):
        updated = queryset.update(status='paid')
        self.message_user(request, f"{updated} invoices marked as paid")
    mark_as_paid.short_description = "Mark selected invoices as paid"
    
    # Export to Excel
    def export_to_xlsx(self, request, queryset):
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"

        # Define headers
        headers = ['ID', 'Customer', 'Invoice Date', 'Due Date', 'Status', 'Total Amount']
        ws.append(headers)

        # Add invoice data
        for invoice in queryset:
            ws.append([
                invoice.id,
                str(invoice.customer),
                invoice.invoice_date.strftime('%Y-%m-%d'),
                invoice.due_date.strftime('%Y-%m-%d'),
                invoice.status,
                float(invoice.total_amount),
            ])

        # Format columns 
        for i, _ in enumerate(headers, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 20

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=invoices.xlsx'
        wb.save(response)
        return response

